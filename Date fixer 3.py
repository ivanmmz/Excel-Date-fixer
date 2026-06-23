"""
Excel Date Fixer Pro - Core Data Processor (CLI Only)
GUI removed. Use Tauri version for the modern UI.
This module retains all processing logic for reference/testing.
"""

import os
import shutil
import warnings
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)


def col_to_index(col_str):
    col_str = col_str.upper().strip()
    if not col_str:
        return None
    if col_str.isdigit():
        return int(col_str) - 1
    index = 0
    for char in col_str:
        if 'A' <= char <= 'Z':
            index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1


class DataProcessor:
    @staticmethod
    def try_parse_any_date(val):
        if pd.isna(val) or val == "" or val is None:
            return None
        if isinstance(val, (datetime, pd.Timestamp)):
            return val if pd.notna(val) else None
        if isinstance(val, (int, float)):
            try:
                return pd.to_datetime(val, unit='D', origin='1899-12-30')
            except Exception:
                return None
        try:
            if isinstance(val, str):
                if "." in val and len(val.split('.')) >= 3:
                    val = val.replace(".", "/")
            res = pd.to_datetime(val, dayfirst=True, errors='coerce')
            return res if pd.notna(res) else None
        except Exception:
            return None

    @staticmethod
    def detect_content_type(dt):
        if dt is None:
            return None
        has_date = dt.year > 1901
        has_time = dt.hour != 0 or dt.minute != 0 or dt.second != 0
        if has_date and has_time:
            return "datetime"
        if has_date:
            return "date"
        return "time"

    @classmethod
    def task_smart_format(cls, wb, log_func, date_fmt, time_fmt, targets):
        dt_fmt = f"{date_fmt} {time_fmt}"
        count = 0
        for sheet in wb.worksheets:
            max_row = sheet.max_row
            if max_row < 2:
                continue
            for row in range(2, max_row + 1):
                for idx, role in targets:
                    if idx is None:
                        continue
                    cell = sheet.cell(row=row, column=idx + 1)
                    dt = cls.try_parse_any_date(cell.value)
                    if dt:
                        if role == "Date":
                            ctype = cls.detect_content_type(dt)
                            cell.value = dt
                            cell.number_format = dt_fmt if ctype == "datetime" else date_fmt
                        else:
                            cell.value = dt.time() if hasattr(dt, 'time') else dt
                            cell.number_format = time_fmt
                        count += 1
        log_func(f"   - Styled {count} target cells across all sheets.")

    @staticmethod
    def task_swap_date(df, d_idx):
        if d_idx is not None and d_idx < len(df.columns):
            col_name = df.columns[d_idx]

            def swap_logic(s):
                if not isinstance(s, str) or ('/' not in s and '-' not in s and '.' not in s):
                    return s
                sep = '/' if '/' in s else ('.' if '.' in s else '-')
                p = s.split(sep)
                if len(p) >= 2:
                    try:
                        p0, p1 = p[0].strip(), p[1].strip().split()[0]
                        if p0.isdigit() and p1.isdigit():
                            if int(p0) <= 31 and int(p1) <= 31:
                                new_s = f"{p1}/{p0}"
                                if len(p) > 2:
                                    new_s += "/" + "/".join(p[2:])
                                return new_s
                    except Exception:
                        pass
                return s

            df[col_name] = df[col_name].astype(str).apply(swap_logic)
            df[col_name] = pd.to_datetime(df[col_name], dayfirst=True, errors='coerce')

    @staticmethod
    def task_fix_missing(df, threshold, protected_indices):
        df_fixed = df.copy()
        for i, col in enumerate(df_fixed.columns):
            if i in protected_indices:
                continue
            is_null = df_fixed[col].isnull()
            if not is_null.any():
                continue
            groups = (is_null != is_null.shift()).cumsum()
            null_groups = groups[is_null]
            group_counts = null_groups.value_counts()
            indices_to_fix = [
                null_groups[null_groups == gid].index.tolist()
                for gid, cnt in group_counts.items()
                if cnt <= threshold
            ]
            for idx_list in indices_to_fix:
                df_fixed.loc[idx_list, col] = df_fixed[col].ffill().loc[idx_list]
        return df_fixed

    @staticmethod
    def task_standardize_1440(df, d_idx, t_idx):
        if d_idx is None or t_idx is None:
            return df
        df_calc = df.copy()
        df_calc['_D_Calc'] = pd.to_datetime(df_calc.iloc[:, d_idx], dayfirst=True, errors='coerce').dt.date
        df_calc['_T_Calc'] = pd.to_datetime(df_calc.iloc[:, t_idx], dayfirst=True, errors='coerce').dt.strftime('%H:%M')
        df_calc = df_calc.dropna(subset=['_D_Calc'])
        result = []
        for _, group in df_calc.groupby('_D_Calc'):
            clean = group.drop_duplicates(subset=['_T_Calc'], keep='first').copy()
            cur = len(clean)
            if cur > 1440:
                clean = clean.iloc[:1440]
            elif cur < 1440:
                clean = pd.concat(
                    [clean, pd.concat([clean.iloc[[-1]]] * (1440 - cur), ignore_index=True)],
                    ignore_index=True
                )
            result.append(clean)
        return pd.concat(result).drop(columns=['_D_Calc', '_T_Calc']) if result else df

    @staticmethod
    def task_date_to_value(df, targets):
        excel_base = pd.Timestamp("1899-12-30")
        df_out = df.copy()
        for idx, role in targets:
            if idx is not None and idx < len(df_out.columns):
                col = df_out.columns[idx]
                dt_s = pd.to_datetime(df_out[col], dayfirst=True, errors='coerce')
                mask = dt_s.notna()
                if role == "Date":
                    df_out.loc[mask, col] = (dt_s[mask] - excel_base) / pd.Timedelta(days=1)
                else:
                    df_out.loc[mask, col] = (dt_s[mask] - dt_s[mask].dt.normalize()).dt.total_seconds() / 86400.0
        return df_out


# ==============================
# CLI Entry Point
# ==============================
if __name__ == "__main__":
    import sys

    def cli_log(msg):
        print(msg)

    if len(sys.argv) < 2:
        print("Usage: python 'Date fixer 3.py' <file.xlsx> [output.xlsx]")
        print("  Or import DataProcessor for programmatic use.")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.dirname(input_path), f"Fixed_{Path(input_path).stem}.xlsx"
    )

    targets = [(0, "Date"), (1, "Time")]
    target_indices = [0, 1]

    print(f">>> Processing: {input_path}")

    sheets_data = pd.read_excel(input_path, sheet_name=None) if not input_path.lower().endswith('.csv') else {
        "Sheet1": pd.read_csv(input_path)
    }
    processed_sheets = {}

    for sheet_name, df in sheets_data.items():
        print(f"  - Sheet: '{sheet_name}' ({len(df)} rows)")
        DataProcessor.task_swap_date(df, 0)
        df = DataProcessor.task_fix_missing(df, 20, target_indices)
        df = DataProcessor.task_standardize_1440(df, 0, 1)
        df = DataProcessor.task_date_to_value(df, targets)
        processed_sheets[sheet_name] = df

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for s_name, s_df in processed_sheets.items():
            s_df.to_excel(writer, sheet_name=s_name, index=False)

    wb = load_workbook(output_path)
    DataProcessor.task_smart_format(wb, cli_log, "DD/MM/YYYY", "HH:mm:ss", targets)
    wb.save(output_path)

    print(f"[✓] SUCCESS: Saved as '{output_path}'")