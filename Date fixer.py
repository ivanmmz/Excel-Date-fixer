import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import ctypes
from datetime import datetime, time, timedelta
from pathlib import Path
import warnings
import shutil

# pyinstaller --noconfirm --onefile --windowed --upx-dir "C:\upx" --icon "D:\Users\Ivan\OneDrive\- Study\Python\App\9. Excel Date fixer\ico.ico" --name "ExcelDateFixerPro" --collect-all customtkinter --collect-all tkinterdnd2 "D:\Users\Ivan\OneDrive\- Study\Python\App\9. Excel Date fixer\Date fixer.py"

# 忽略不必要的警告以保持控制台整洁
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)

# 尝试加载拖拽库
HAS_TKDND = False
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD

    HAS_TKDND = True
except ImportError:
    pass

HAS_WINDND = False
try:
    import windnd

    HAS_WINDND = True
except ImportError:
    pass

# ==============================
# Windows 11 Mica 效果适配
# ==============================
try:
    from ctypes import windll, byref, sizeof, c_int


    def apply_mica_style(window):
        window.update()
        hwnd = windll.user32.GetParent(window.winfo_id())
        value = c_int(2)
        windll.dwmapi.DwmSetWindowAttribute(hwnd, 20, byref(value), sizeof(value))
        backdrop = c_int(2)
        windll.dwmapi.DwmSetWindowAttribute(hwnd, 38, byref(backdrop), sizeof(backdrop))
except:
    def apply_mica_style(window):
        pass


# ==============================
# 辅助函数
# ==============================
def col_to_index(col_str):
    """将 Excel 列字母（A, B...）转换为从 0 开始的索引。"""
    col_str = col_str.upper().strip()
    if not col_str: return None
    if col_str.isdigit(): return int(col_str) - 1
    index = 0
    for char in col_str:
        if 'A' <= char <= 'Z': index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1


# ==============================
# 核心逻辑：数据处理器
# ==============================
class DataProcessor:
    @staticmethod
    def try_parse_any_date(val):
        """稳健地解析包含字符串、Datetime 以及 Excel 序列号的值。"""
        if pd.isna(val) or val == "" or val is None: return None
        if isinstance(val, (datetime, pd.Timestamp)):
            return val if pd.notna(val) else None
        if isinstance(val, (int, float)):
            try:
                return pd.to_datetime(val, unit='D', origin='1899-12-30')
            except:
                return None
        try:
            if isinstance(val, str):
                if "." in val and len(val.split('.')) >= 3:
                    val = val.replace(".", "/")
            res = pd.to_datetime(val, dayfirst=True, errors='coerce')
            return res if pd.notna(res) else None
        except:
            return None

    @staticmethod
    def detect_content_type(dt):
        """分析 Datetime 对象，判断其内容倾向。"""
        if dt is None: return None
        has_date = dt.year > 1901
        has_time = dt.hour != 0 or dt.minute != 0 or dt.second != 0
        if has_date and has_time: return "datetime"
        if has_date: return "date"
        return "time"

    @classmethod
    def task_smart_format(cls, wb, log_func, date_fmt, time_fmt, targets):
        """任务 4：应用视觉格式。严格遵循用户定义的列角色（Date/Time）。"""
        log_func(f"Task: Strictly applying user formats based on column roles...")
        dt_fmt = f"{date_fmt} {time_fmt}"

        for sheet in wb.worksheets:
            max_row = sheet.max_row
            if max_row < 2: continue
            for row in range(2, max_row + 1):
                for idx, role in targets:
                    if idx is None: continue
                    cell = sheet.cell(row=row, column=idx + 1)
                    dt = cls.try_parse_any_date(cell.value)
                    if dt:
                        if role == "Date":
                            ctype = cls.detect_content_type(dt)
                            cell.value = dt
                            cell.number_format = dt_fmt if ctype == "datetime" else date_fmt
                        else:  # Time role: Forced strip date
                            cell.value = dt.time() if hasattr(dt, 'time') else dt
                            cell.number_format = time_fmt

    @staticmethod
    def task_swap_date(df, d_idx):
        """任务 1：针对主日期列执行日月对调。"""
        if d_idx is not None and d_idx < len(df.columns):
            col_name = df.columns[d_idx]

            def swap_logic(s):
                if not isinstance(s, str) or ('/' not in s and '-' not in s and '.' not in s): return s
                sep = '/' if '/' in s else ('.' if '.' in s else '-')
                p = s.split(sep)
                if len(p) >= 2:
                    try:
                        if p[0].strip().isdigit() and p[1].strip().split()[0].isdigit():
                            parts_0 = p[0].strip()
                            parts_1 = p[1].strip().split()[0]
                            if int(parts_0) <= 31 and int(parts_1) <= 31:
                                new_s = f"{parts_1}/{parts_0}"
                                if len(p) > 2: new_s += "/" + "/".join(p[2:])
                                return new_s
                    except:
                        pass
                return s

            df[col_name] = df[col_name].astype(str).apply(swap_logic)
            df[col_name] = pd.to_datetime(df[col_name], dayfirst=True, errors='coerce')

    @staticmethod
    def task_fix_missing(df, threshold, protected_indices):
        """任务 2：补全数值缺失，避开日期时间列。"""
        df_fixed = df.copy()
        for i, col in enumerate(df_fixed.columns):
            if i in protected_indices: continue
            is_null = df_fixed[col].isnull()
            if not is_null.any(): continue
            groups = (is_null != is_null.shift()).cumsum()
            null_groups = groups[is_null]
            group_counts = null_groups.value_counts()
            indices_to_fix = [null_groups[null_groups == gid].index.tolist() for gid, cnt in group_counts.items() if
                              cnt <= threshold]
            for idx_list in indices_to_fix:
                df_fixed.loc[idx_list, col] = df_fixed[col].ffill().loc[idx_list]
        return df_fixed

    @staticmethod
    def task_standardize_1440(df, d_idx, t_idx):
        """任务 3：1440 行标准化。"""
        if d_idx is None or t_idx is None: return df
        df_calc = df.copy()
        df_calc['_D_Calc'] = pd.to_datetime(df_calc.iloc[:, d_idx], dayfirst=True, errors='coerce').dt.date
        df_calc['_T_Calc'] = pd.to_datetime(df_calc.iloc[:, t_idx], dayfirst=True, errors='coerce').dt.strftime('%H:%M')
        df_calc = df_calc.dropna(subset=['_D_Calc'])
        result = []
        for _, group in df_calc.groupby('_D_Calc'):
            clean = group.drop_duplicates(subset=['_T_Calc'], keep='first').copy()
            cur = len(clean)
            if cur > 1440:
                clean = clean.iloc[:1440]
            elif cur < 1440:
                clean = pd.concat([clean, pd.concat([clean.iloc[[-1]]] * (1440 - cur), ignore_index=True)],
                                  ignore_index=True)
            result.append(clean)
        return pd.concat(result).drop(columns=['_D_Calc', '_T_Calc']) if result else df

    @staticmethod
    def task_date_to_value(df, targets):
        """任务 5：转换为 Excel 序列号。尊重用户定义的角色。"""
        excel_base = pd.Timestamp("1899-12-30")
        df_out = df.copy()
        for idx, role in targets:
            if idx is not None and idx < len(df_out.columns):
                col = df_out.columns[idx]
                dt_s = pd.to_datetime(df_out[col], dayfirst=True, errors='coerce')
                mask = dt_s.notna()
                if role == "Date":
                    df_out.loc[mask, col] = (dt_s[mask] - excel_base) / pd.Timedelta(days=1)
                else:  # Time role
                    df_out.loc[mask, col] = (dt_s[mask] - dt_s[mask].dt.normalize()).dt.total_seconds() / 86400.0
        return df_out


# ==============================
# GUI 应用程序类
# ==============================
class UnifiedApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        if HAS_TKDND:
            try:
                self.TkDnDVersion = self.tk.call('TkinterDnD_Init')
            except:
                pass

        self.title("Excel Data Fixer Suite Pro")
        self.geometry("1150x950")
        self.minsize(1100, 800)

        self.file_list = []
        self.task_queue = []
        self.missing_threshold = tk.IntVar(value=20)

        # 列定义变量
        self.col1_str = tk.StringVar(value="A")
        self.col1_role = tk.StringVar(value="Date")

        self.col2_str = tk.StringVar(value="B")
        self.col2_role = tk.StringVar(value="Time")

        self.col3_str = tk.StringVar(value="")
        self.col3_role = tk.StringVar(value="Date")

        self.date_format_var = tk.StringVar(value="dd/mm/yyyy")
        self.time_format_var = tk.StringVar(value="hh:mm")
        self.log_font_size = 12

        self.setup_ui()
        apply_mica_style(self)
        self.init_drag_and_drop_logic()

    def setup_ui(self):
        ctk.set_appearance_mode("dark")
        self.grid_columnconfigure(0, weight=4)
        self.grid_columnconfigure(1, weight=6)
        self.grid_rowconfigure(0, weight=1)

        # --- 左侧面板 ---
        self.left_panel = ctk.CTkScrollableFrame(self, label_text="Step Configuration")
        self.left_panel.grid(row=0, column=0, padx=(20, 10), pady=20, sticky="nsew")

        # 列定义卡片
        global_card = ctk.CTkFrame(self.left_panel, fg_color="#3d3d3d", corner_radius=10)
        global_card.pack(fill="x", padx=10, pady=5)

        ctk.CTkLabel(global_card, text="Define Target Columns & Roles:", font=ctk.CTkFont(weight="bold")).grid(row=0,
                                                                                                               column=0,
                                                                                                               columnspan=3,
                                                                                                               padx=10,
                                                                                                               pady=10,
                                                                                                               sticky="w")

        # Column 1
        ctk.CTkEntry(global_card, textvariable=self.col1_str, width=50).grid(row=1, column=0, padx=(15, 5), pady=5)
        ctk.CTkSegmentedButton(global_card, values=["Date", "Time"], variable=self.col1_role).grid(row=1, column=1,
                                                                                                   padx=5, pady=5)
        ctk.CTkLabel(global_card, text="(Primary)", text_color="gray").grid(row=1, column=2, padx=5, pady=5, sticky="w")

        # Column 2
        ctk.CTkEntry(global_card, textvariable=self.col2_str, width=50).grid(row=2, column=0, padx=(15, 5), pady=5)
        ctk.CTkSegmentedButton(global_card, values=["Date", "Time"], variable=self.col2_role).grid(row=2, column=1,
                                                                                                   padx=5, pady=5)

        # Column 3
        ctk.CTkEntry(global_card, textvariable=self.col3_str, width=50).grid(row=3, column=0, padx=(15, 5), pady=10)
        ctk.CTkSegmentedButton(global_card, values=["Date", "Time"], variable=self.col3_role).grid(row=3, column=1,
                                                                                                   padx=5, pady=10)

        # 任务选项
        self.create_task_card(self.left_panel, "1. Swap Day/Month", "swap_date",
                              "Corrects swapped DD/MM in Primary Col.")
        self.create_task_card(self.left_panel, "2. Gap Filling", "fix_missing", "Fill numeric gaps in non-target cols.",
                              has_param=True)
        self.create_task_card(self.left_panel, "3. 1440 Standardization", "standard_1440",
                              "Align rows (Destructive to styles).")
        self.create_task_card(self.left_panel, "4. Visual Styling", "smart_format", "STRICTLY apply user formats.",
                              has_formats=True)
        self.create_task_card(self.left_panel, "5. Date to Value", "date_to_value",
                              "Convert defined cols to Excel numbers.")

        self.queue_frame = ctk.CTkFrame(self.left_panel, fg_color="#2b2b2b")
        self.queue_frame.pack(fill="x", padx=10, pady=20)
        ctk.CTkLabel(self.queue_frame, text="Execution Queue:", font=ctk.CTkFont(weight="bold")).pack(pady=5)
        self.queue_label = ctk.CTkLabel(self.queue_frame, text="(Select tasks to start)", text_color="#888",
                                        wraplength=350)
        self.queue_label.pack(pady=5)

        # --- 右侧面板 ---
        self.right_panel = ctk.CTkFrame(self, fg_color="transparent")
        self.right_panel.grid(row=0, column=1, padx=(10, 20), pady=20, sticky="nsew")
        self.list_label = ctk.CTkLabel(self.right_panel, text="File Queue:", font=ctk.CTkFont(weight="bold"))
        self.list_label.pack(anchor="w", pady=(0, 5))
        self.file_scroll_frame = ctk.CTkScrollableFrame(self.right_panel, height=200, fg_color="#1a1a1a")
        self.file_scroll_frame.pack(fill="x", pady=(0, 10))
        self.log_label = ctk.CTkLabel(self.right_panel, text="Log (Ctrl+Scroll to Zoom):",
                                      font=ctk.CTkFont(weight="bold"))
        self.log_label.pack(anchor="w", pady=(10, 5))
        self.log_display = ctk.CTkTextbox(self.right_panel, font=("Consolas", self.log_font_size), text_color="#00ff00",
                                          fg_color="#111")
        self.log_display.pack(fill="both", expand=True)
        self.log_display.bind("<Control-MouseWheel>", self.handle_zoom)

        # --- 底部控制 ---
        self.bottom_frame = ctk.CTkFrame(self, height=100)
        self.bottom_frame.grid(row=1, column=0, columnspan=2, padx=20, pady=(0, 20), sticky="ew")
        self.prog_bar = ctk.CTkProgressBar(self.bottom_frame)
        self.prog_bar.pack(fill="x", padx=25, pady=(15, 5))
        self.prog_bar.set(0)
        btn_box = ctk.CTkFrame(self.bottom_frame, fg_color="transparent")
        btn_box.pack(fill="x", padx=20, pady=(0, 10))
        ctk.CTkButton(btn_box, text="Add Files", command=self.add_files_manually, width=120).pack(side="left", padx=5)
        ctk.CTkButton(btn_box, text="Clear", fg_color="#444", command=self.clear_files_list, width=80).pack(side="left",
                                                                                                            padx=5)
        self.run_btn = ctk.CTkButton(btn_box, text="Run Pipeline", fg_color="#1f6aa5",
                                     command=self.start_processing_thread, width=180, font=ctk.CTkFont(weight="bold"))
        self.run_btn.pack(side="right", padx=5)

    def init_drag_and_drop_logic(self):
        if HAS_WINDND:
            windnd.hook_dropfiles(self, lambda files: self.after(10, self._on_drop_safe, files))
        elif HAS_TKDND:
            self.file_scroll_frame.drop_target_register(DND_FILES)
            self.file_scroll_frame.dnd_bind('<<Drop>>',
                                            lambda e: self.after(10, self._on_drop_safe, self.tk.splitlist(e.data)))

    def _on_drop_safe(self, files):
        self._append_files([f.decode('gbk') if isinstance(f, bytes) else f for f in files])

    def handle_zoom(self, event):
        self.log_font_size = max(8, min(50, self.log_font_size + (1 if event.delta > 0 else -1)))
        self.log_display.configure(font=("Consolas", self.log_font_size))

    def create_task_card(self, master, title, task_id, desc, has_param=False, has_formats=False):
        card = ctk.CTkFrame(master, fg_color="#333", corner_radius=10)
        card.pack(fill="x", padx=10, pady=5)
        row1 = ctk.CTkFrame(card, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=(10, 5))
        var = tk.BooleanVar(value=False)
        cb = ctk.CTkCheckBox(row1, text=title, font=ctk.CTkFont(size=14, weight="bold"), variable=var,
                             command=lambda: self.update_task_queue(task_id, var.get()))
        cb.pack(side="left")
        if has_param:
            p_ui = ctk.CTkFrame(row1, fg_color="transparent")
            p_ui.pack(side="right")
            ctk.CTkEntry(p_ui, textvariable=self.missing_threshold, width=40).pack(side="left", padx=5)
            ctk.CTkLabel(p_ui, text="pts").pack(side="left")
        if has_formats:
            f_ui = ctk.CTkFrame(card, fg_color="transparent")
            f_ui.pack(fill="x", padx=15, pady=(0, 5))
            ctk.CTkLabel(f_ui, text="Date Fmt:").grid(row=0, column=0, sticky="w")
            ctk.CTkEntry(f_ui, textvariable=self.date_format_var, width=100, height=22).grid(row=0, column=1, padx=5,
                                                                                             pady=2)
            ctk.CTkLabel(f_ui, text="Time Fmt:").grid(row=1, column=0, sticky="w")
            ctk.CTkEntry(f_ui, textvariable=self.time_format_var, width=100, height=22).grid(row=1, column=1, padx=5,
                                                                                             pady=2)
        ctk.CTkLabel(card, text=desc, font=ctk.CTkFont(size=11), text_color="#bbb").pack(anchor="w", padx=15,
                                                                                         pady=(0, 10))

    def update_task_queue(self, task_id, is_selected):
        if is_selected:
            if task_id not in self.task_queue: self.task_queue.append(task_id)
        elif task_id in self.task_queue:
            self.task_queue.remove(task_id)
        names = {"swap_date": "Swap", "fix_missing": "Gap Fill", "standard_1440": "1440 Std", "smart_format": "Style",
                 "date_to_value": "Dt->Val"}
        self.queue_label.configure(text=" → ".join([f"[{i + 1}] {names[tid]}" for i, tid in enumerate(
            self.task_queue)]) if self.task_queue else "(Select tasks)")

    def _append_files(self, files):
        for f in [f for f in files if f.lower().endswith(('.xlsx', '.xls', '.csv'))]:
            if f not in self.file_list: self.file_list.append(f)
        self.refresh_file_list_ui()

    def refresh_file_list_ui(self):
        for widget in self.file_scroll_frame.winfo_children(): widget.destroy()
        for f in self.file_list:
            row = ctk.CTkFrame(self.file_scroll_frame, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=f"• {os.path.basename(f)}", anchor="w").pack(side="left", fill="x", expand=True,
                                                                                padx=5)
            ctk.CTkButton(row, text="✕", width=25, height=25, fg_color="#b22222",
                          command=lambda p=f: self.remove_file(p)).pack(side="right", padx=5)

    def remove_file(self, path):
        if path in self.file_list: self.file_list.remove(path)
        self.refresh_file_list_ui()

    def add_to_log(self, message):
        ts = datetime.now().strftime('%H:%M:%S')
        self.after(10, self._append_log_text, f"[{ts}] {message}\n")

    def _append_log_text(self, msg):
        self.log_display.configure(state="normal")
        self.log_display.insert("end", msg)
        self.log_display.see("end")

    def add_files_manually(self):
        self._append_files(filedialog.askopenfilenames(filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")]))

    def clear_files_list(self):
        (self.file_list.clear(), self.refresh_file_list_ui(), self.prog_bar.set(0))

    def start_processing_thread(self):
        if not self.file_list or not self.task_queue: return messagebox.showwarning("Warning", "Invalid input.")
        self.run_btn.configure(state="disabled")
        threading.Thread(target=self.run_processor, daemon=True).start()

    def run_processor(self):
        total = len(self.file_list)
        # 准备目标列表
        raw_targets = [
            (col_to_index(self.col1_str.get()), self.col1_role.get()),
            (col_to_index(self.col2_str.get()), self.col2_role.get()),
            (col_to_index(self.col3_str.get()), self.col3_role.get())
        ]
        targets = [(idx, role) for idx, role in raw_targets if idx is not None]
        target_indices = [idx for idx, _ in targets]
        d_idx = target_indices[0] if target_indices else None
        t_idx = target_indices[1] if len(target_indices) > 1 else None
        use_std_1440 = "standard_1440" in self.task_queue

        for i, path in enumerate(self.file_list):
            try:
                name = os.path.basename(path)
                self.add_to_log(f"Processing: {name}")
                output = os.path.join(os.path.dirname(path), f"Fixed_{Path(path).stem}.xlsx")
                sheets_data = pd.read_excel(path, sheet_name=None) if not path.lower().endswith('.csv') else {
                    "Sheet1": pd.read_csv(path)}
                processed_sheets = {}

                for sheet_name, df in sheets_data.items():
                    for task_id in self.task_queue:
                        if task_id == "swap_date":
                            DataProcessor.task_swap_date(df, d_idx)
                        elif task_id == "fix_missing":
                            df = DataProcessor.task_fix_missing(df, self.missing_threshold.get(), target_indices)
                        elif task_id == "standard_1440":
                            df = DataProcessor.task_standardize_1440(df, d_idx, t_idx)
                        elif task_id == "date_to_value":
                            df = DataProcessor.task_date_to_value(df, targets)
                    for idx, col in enumerate(df.columns):
                        if idx not in target_indices:
                            try:
                                df[col] = pd.to_numeric(df[col])
                            except:
                                pass
                    processed_sheets[sheet_name] = df

                if use_std_1440 or path.lower().endswith('.csv'):
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        for s_name, s_df in processed_sheets.items(): s_df.to_excel(writer, sheet_name=s_name,
                                                                                    index=False)
                else:
                    shutil.copyfile(path, output)
                    wb = load_workbook(output)
                    for s_name, s_df in processed_sheets.items():
                        if s_name in wb.sheetnames:
                            ws = wb[s_name]
                            for r_idx, row_vals in enumerate(s_df.values):
                                for c_idx, val in enumerate(row_vals):
                                    cell = ws.cell(row=r_idx + 2, column=c_idx + 1)
                                    # 智能值注入：如果该列被定义为 Time 角色且内容包含日期，剥离日期
                                    is_time_col = False
                                    for t_i, t_role in targets:
                                        if c_idx == t_i and t_role == "Time":
                                            is_time_col = True
                                            break

                                    if is_time_col:
                                        dt = DataProcessor.try_parse_any_date(val)
                                        cell.value = dt.time() if dt and hasattr(dt, 'time') else val
                                    else:
                                        cell.value = val
                    wb.save(output)

                if "smart_format" in self.task_queue:
                    wb = load_workbook(output)
                    DataProcessor.task_smart_format(wb, self.add_to_log, self.date_format_var.get(),
                                                    self.time_format_var.get(), targets)
                    wb.save(output)

                self.add_to_log(f"Finished: {os.path.basename(output)}")
                self.after(10, lambda v=(i + 1) / total: self.prog_bar.set(v))
            except Exception as e:
                self.add_to_log(f"!!! Error [{name}]: {str(e)}")

        self.after(100, lambda: (self.run_btn.configure(state="normal"), messagebox.showinfo("Success", "Done!")))


if __name__ == "__main__":
    app = UnifiedApp()
    app.mainloop()